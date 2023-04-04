# -*- coding: utf-8 -*-

"""
Created on Sun Apr  2 10:05:50 2023

@author: Mayk Al-Ghrawi
"""

import openpyxl
import datetime


class Session:
    def __init__(self, date, start_time, end_time, course_id, trainer_name, trainee_attendance):
        self.date = date
        self.start_time = start_time
        self.end_time = end_time
        self.course_id = course_id
        self.trainer_name = trainer_name
        self.trainee_attendance = trainee_attendance
class Manager:
    def __init__(self, manager_id, full_name, email_id, phone_number):
        self.manager_id = manager_id
        self.full_name = full_name
        self.email_id = email_id
        self.phone_number = phone_number

class Trainee:
    def __init__(self, trainee_id, name, course, background, work_experience):
        self.trainee_id = trainee_id
        self.name = name
        self.course = course
        self.background = background
        self.work_experience = work_experience
class Trainer:
    def __init__(self, trainer_id, full_name, email_id, phone_number):
        self.trainer_id = trainer_id
        self.full_name = full_name
        self.email_id = email_id
        self.phone_number = phone_number

class TraineeManager:
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.Workbook()

        # Trainee sheet setup
        self.trainee_sheet = self.wb.active
        self.trainee_sheet.title = "ListOfTrainees"
        self.create_headers(self.trainee_sheet, ['ID', 'Name', 'Course', 'Background', 'Work Experience'])
    def create_headers(self, sheet, headers):
        for idx, header in enumerate(headers, start=1):
            try:
                sheet.cell(row=1, column=idx, value=header)
            except ValueError as e:
                print(f"Error: {e}")


    def add_course_details(self, course_id, course_description):
        if "CourseDetails" not in self.wb.sheetnames:
            course_sheet = self.wb.create_sheet("CourseDetails")
            self.create_headers(course_sheet, ['Course ID', 'Course Description'])
        else:
            course_sheet = self.wb["CourseDetails"]

        row = course_sheet.max_row + 1
        course_sheet.cell(row=row, column=1, value=course_id)
        course_sheet.cell(row=row, column=2, value=course_description)
    def add_trainee(self, trainee):
        row = self.trainee_sheet.max_row + 1
        self.trainee_sheet.cell(row=row, column=1, value=trainee.trainee_id)
        self.trainee_sheet.cell(row=row, column=2, value=trainee.name)
        self.trainee_sheet.cell(row=row, column=3, value=trainee.course)
        self.trainee_sheet.cell(row=row, column=4, value=trainee.background)
        self.trainee_sheet.cell(row=row, column=5, value=trainee.work_experience)


    def delete_trainee(self, trainee_id):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == trainee_id:
                self.sheet.delete_rows(row[0].row)
                break

    def update_trainee(self, trainee):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == trainee.trainee_id:
                row[1].value = trainee.name
                row[2].value = trainee.course
                row[3].value = trainee.background
                row[4].value = trainee.work_experience
                break

    def save(self):
        self.wb.save(self.filename)
        
    def add_trainer_details(self, trainer):
        if "TrainerDetails" not in self.wb.sheetnames:
            trainer_sheet = self.wb.create_sheet("TrainerDetails")
            self.create_headers(trainer_sheet, ['Trainer ID', 'Full Name', 'Email ID', 'Phone Number'])
        else:
            trainer_sheet = self.wb["TrainerDetails"]

        row = trainer_sheet.max_row + 1
        trainer_sheet.cell(row=row, column=1, value=trainer.trainer_id)
        trainer_sheet.cell(row=row, column=2, value=trainer.full_name)
        trainer_sheet.cell(row=row, column=3, value=trainer.email_id)
        trainer_sheet.cell(row=row, column=4, value=trainer.phone_number)

    def mapping_course_trainer(self, course_id, trainer_id):
        if "CourseTrainerMapping" not in self.wb.sheetnames:
            mapping_sheet = self.wb.create_sheet("CourseTrainerMapping")
            self.create_headers(mapping_sheet, ['Course ID', 'Trainer ID'])
        else:
            mapping_sheet = self.wb["CourseTrainerMapping"]

        row = mapping_sheet.max_row + 1
        mapping_sheet.cell(row=row, column=1, value=course_id)
        mapping_sheet.cell(row=row, column=2, value=trainer_id)
    def add_manager_details(self, manager):
        if "ManagerDetails" not in self.wb.sheetnames:
            manager_sheet = self.wb.create_sheet("ManagerDetails")
            self.create_headers(manager_sheet, ['Manager ID', 'Full Name', 'Email ID', 'Phone Number'])
        else:
            manager_sheet = self.wb["ManagerDetails"]

        row = manager_sheet.max_row + 1
        manager_sheet.cell(row=row, column=1, value=manager.manager_id)
        manager_sheet.cell(row=row, column=2, value=manager.full_name)
        manager_sheet.cell(row=row, column=3, value=manager.email_id)
        manager_sheet.cell(row=row, column=4, value=manager.phone_number)
    def add_session(self, session):
        session_date = session.date.strftime("%B%d_%Y")
        session_sheet_name = f"{session_date}"
        
        if session_sheet_name not in self.wb.sheetnames:
            session_sheet = self.wb.create_sheet(session_sheet_name)
            self.create_headers(session_sheet, ['Date', 'Start Time', 'End Time', 'Course ID', 'Trainer Name', 'Trainee ID', 'Trainee Name', 'Attendance'])
        else:
            session_sheet = self.wb[session_sheet_name]

        for trainee_id, attendance in session.trainee_attendance.items():
            row = session_sheet.max_row + 1
            session_sheet.cell(row=row, column=1, value=session.date)
            session_sheet.cell(row=row, column=2, value=session.start_time)
            session_sheet.cell(row=row, column=3, value=session.end_time)
            session_sheet.cell(row=row, column=4, value=session.course_id)
            session_sheet.cell(row=row, column=5, value=session.trainer_name)
            session_sheet.cell(row=row, column=6, value=trainee_id)
            session_sheet.cell(row=row, column=7, value=self.get_trainee_name(trainee_id))
            session_sheet.cell(row=row, column=8, value=attendance)

    def get_trainee_name(self, trainee_id):
        for row in self.trainee_sheet.iter_rows(min_row=2):
            if row[0].value == trainee_id:
                return row[1].value
        return None
    
def print_menu():
    print("1. Add Trainee")
    print("2. Delete Trainee")
    print("3. Update Trainee")
    print("4. Add Course Details")
    print("5. Add Trainer Details")
    print("6. Map Course to Trainer")
    print("7. Add Manager Details")
    print("8. Add Session")
    print("9. Save and Exit")
    print("0. Exit without saving")

def main_menu():
    manager = TraineeManager("trainees_courses_trainers_managers_sessions.xlsx")
    while True:
        print_menu()
        try:
            choice = int(input("Enter your choice: "))
            if choice == 1:
                trainee_id = int(input("Enter Trainee ID: "))
                name = input("Enter Trainee Name: ")
                course = input("Enter Course: ")
                background = input("Enter Background/Degree: ")
                work_experience = int(input("Enter Work Experience (in years): "))
                trainee = Trainee(trainee_id, name, course, background, work_experience)
                manager.add_trainee(trainee)
            elif choice == 2:
                trainee_id = int(input("Enter Trainee ID: "))
                manager.delete_trainee(trainee_id)
            elif choice == 3:
                trainee_id = int(input("Enter Trainee ID: "))
                new_name = input("Enter New Trainee Name: ")
                new_course = input("Enter New Course: ")
                new_background = input("Enter New Background/Degree: ")
                new_work_experience = int(input("Enter New Work Experience (in years): "))
                updated_trainee = Trainee(trainee_id, new_name, new_course, new_background, new_work_experience)
                manager.update_trainee(updated_trainee)
            elif choice == 4:
                course_id = input("Enter Course ID: ")
                description = input("Enter Course Description: ")
                manager.add_course_details(course_id, description)
            elif choice == 5:
                trainer_id = int(input("Enter Trainer ID: "))
                full_name = input("Enter Trainer Full Name: ")
                email_id = input("Enter Trainer Email ID: ")
                phone_number = input("Enter Trainer Phone Number: ")
                trainer = Trainer(trainer_id, full_name, email_id, phone_number)
                manager.add_trainer_details(trainer)
            elif choice == 6:
                course_id = input("Enter Course ID: ")
                trainer_id = int(input("Enter Trainer ID: "))
                manager.mapping_course_trainer(course_id, trainer_id)
            elif choice == 7:
                manager_id = int(input("Enter Manager ID: "))
                full_name = input("Enter Manager Full Name: ")
                email_id = input("Enter Manager Email ID: ")
                phone_number = input("Enter Manager Phone Number: ")
                manager_obj = Manager(manager_id, full_name, email_id, phone_number)
                manager.add_manager_details(manager_obj)
            elif choice == 8:
                session_date_input = input("Enter session date (YYYY-MM-DD): ")
                try:
                    session_date = datetime.datetime.strptime(session_date_input, "%Y-%m-%d").date()
                except ValueError:
                    print("Invalid date format. Please use the format YYYY-MM-DD.")
                    continue
                start_time = input("Enter session start time (e.g., 9:00 AM): ")
                end_time = input("Enter session end time (e.g., 5:00 PM): ")
                course_id = input("Enter Course ID: ")
                trainer_name = input("Enter Trainer Name: ")
                num_trainees = int(input("Enter the number of trainees: "))
                trainee_attendance = {}
                for i in range(num_trainees):
                    while True:
                        trainee_id = int(input(f"Enter Trainee {i + 1} ID: "))
                        trainee_name = manager.get_trainee_name(trainee_id)
                        if trainee_name is None:
                            print("Invalid Trainee ID. Please enter a valid ID.")
                        else:
                            break
                    while True:
                        attendance = input(f"Enter Trainee {i + 1} Attendance (P/A): ")
                        if attendance.upper() not in ["P", "A"]:
                            print("Invalid attendance value. Please enter 'P' or 'A'.")
                        else:
                            break
                    trainee_attendance[trainee_id] = attendance.upper()
                
                session = Session(session_date, start_time, end_time, course_id, trainer_name, trainee_attendance)
                manager.add_session(session)
        except ValueError:
            print("Invalid input. Please enter a number.")

        if choice == 9:
            manager.save()
            print("Data saved. Exiting...")
            break
        elif choice == 0:
            print("Exiting without saving...")
            break

if __name__ == "__main__":
    main_menu()